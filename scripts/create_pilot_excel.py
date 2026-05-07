from __future__ import annotations

import argparse
import csv
import json
import math
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from visum.numbers import extract_numbers
from visum.text import sentence_split, word_count


IMPORTANT_RE = re.compile(
    r"(VN-?Index|HNX-?Index|UPCoM|VN30|HoSE|HNX|thanh khoản|giá trị giao dịch|"
    r"khối ngoại|mua ròng|bán ròng|cổ phiếu|đồng/cp|tỷ đồng|triệu cổ phiếu|"
    r"doanh thu|lợi nhuận|lãi|lỗ|cổ tức|tăng|giảm|điểm|%)",
    flags=re.IGNORECASE,
)

STOCK_RE = re.compile(
    r"(VN-?Index|HNX-?Index|UPCoM|VN30|HoSE|HNX|chứng khoán|cổ phiếu|"
    r"thanh khoản|khối ngoại|mua ròng|bán ròng|tự doanh|"
    r"cổ tức|đồng/cp|niêm yết|đại hội cổ đông|ĐHĐCĐ|sàn giao dịch|"
    r"bluechip|vốn hóa|vốn hoá)",
    flags=re.IGNORECASE,
)


def read_csv(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        return list(csv.DictReader(file))


def difficulty_from_number_count(count: int) -> str:
    if count <= 20:
        return "Easy"
    if count <= 35:
        return "Medium"
    if count <= 50:
        return "Hard"
    return "Extreme"


def unique_keep_order(items: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for item in items:
        key = item.lower().replace(" ", "")
        if key in seen:
            continue
        seen.add(key)
        out.append(item)
    return out


def infer_required_numbers(row: dict, max_required: int = 45) -> list[str]:
    title = row.get("title", "")
    text = row.get("text", "")
    pieces: list[str] = [title]

    sentences = sentence_split(text)
    pieces.extend(sentences[:3])
    pieces.extend(sentence for sentence in sentences if IMPORTANT_RE.search(sentence))

    required: list[str] = []
    for piece in pieces:
        required.extend(extract_numbers(piece))
    required = unique_keep_order(required)
    if len(required) < 5:
        for number in extract_numbers(text):
            if number not in required:
                required.append(number)
            if len(required) >= 5:
                break
    return required[:max_required]


def is_stock_relevant(row: dict) -> bool:
    haystack = " ".join([row.get("title", ""), row.get("text", "")])
    return bool(STOCK_RE.search(haystack))


def prepare_rows(rows: list[dict], target: int, require_stock_keyword: bool = True) -> list[dict]:
    prepared: list[dict] = []
    for row in rows:
        if require_stock_keyword and not is_stock_relevant(row):
            continue
        text = row.get("text", "")
        numbers = extract_numbers(text)
        required = infer_required_numbers(row)
        source_words = word_count(text)
        item = {
            **row,
            "source_word_count": source_words,
            "min_summary_words": math.ceil(source_words * 0.20),
            "max_summary_words": math.floor(source_words * 0.25),
            "all_numbers": "; ".join(numbers),
            "number_count": len(numbers),
            "required_numbers": "; ".join(required),
            "required_number_count": len(required),
            "difficulty": difficulty_from_number_count(len(numbers)),
            "summary_word_count": "",
            "length_ratio": "",
            "length_compliance": "",
            "missing_numbers": "",
            "hallucinated_numbers": "",
            "number_accuracy": "",
            "validation_mode": "all_numbers",
            "status": "Not Started",
            "reject_reason": "",
            "annotator_note": "",
        }
        prepared.append(item)

    rank = {"Easy": 0, "Medium": 1, "Hard": 2, "Extreme": 3}
    prepared.sort(key=lambda item: (rank[item["difficulty"]], item["number_count"], item["source_word_count"]))
    return prepared[:target]


def write_workbook(rows: list[dict], output: Path) -> None:
    columns = [
        "id",
        "url",
        "title",
        "source",
        "published_at",
        "text",
        "source_word_count",
        "min_summary_words",
        "max_summary_words",
        "all_numbers",
        "number_count",
        "required_numbers",
        "required_number_count",
        "difficulty",
        "summary",
        "summary_word_count",
        "length_ratio",
        "length_compliance",
        "missing_numbers",
        "hallucinated_numbers",
        "number_accuracy",
        "validation_mode",
        "status",
        "reject_reason",
        "annotator_note",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Annotate"
    ws.append(columns)
    for row in rows:
        ws.append([row.get(column, "") for column in columns])

    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D1D5DB")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)

    widths = {
        "A": 18,
        "B": 42,
        "C": 48,
        "D": 24,
        "E": 18,
        "F": 72,
        "G": 16,
        "H": 18,
        "I": 18,
        "J": 48,
        "K": 14,
        "L": 48,
        "M": 22,
        "N": 14,
        "O": 72,
        "P": 18,
        "Q": 14,
        "R": 18,
        "S": 42,
        "T": 42,
        "U": 16,
        "V": 18,
        "W": 16,
        "X": 24,
        "Y": 36,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    status_dv = DataValidation(type="list", formula1='"Not Started,Valid,Revise,Reject"', allow_blank=True)
    reason_dv = DataValidation(
        type="list",
        formula1='"Too many numbers,Too short,Too long,Too noisy,Not stock market,Duplicate,Other"',
        allow_blank=True,
    )
    mode_dv = DataValidation(type="list", formula1='"all_numbers,required_numbers"', allow_blank=True)
    ws.add_data_validation(status_dv)
    ws.add_data_validation(reason_dv)
    ws.add_data_validation(mode_dv)
    last_row = max(ws.max_row, 2)
    status_dv.add(f"W2:W{last_row}")
    reason_dv.add(f"X2:X{last_row}")
    mode_dv.add(f"V2:V{last_row}")

    guide = wb.create_sheet("Guideline")
    guide_rows = [
        ["Mục tiêu", "Summary dài 20-25% số từ của text gốc."],
        ["Số liệu", "Mặc định validation_mode=all_numbers: giữ toàn bộ all_numbers."],
        ["Chế độ nhẹ hơn", "Có thể đổi validation_mode=required_numbers nếu bài quá nhiều số, nhưng phải ghi rõ trong báo cáo."],
        ["Không được", "Không làm tròn số, không đổi đơn vị, không thêm số mới, không đảo tăng/giảm."],
        ["Reject", "Reject nếu bài quá nhiều số, quá nhiễu, không thuộc chứng khoán, hoặc không thể tóm tắt trong 20-25%."],
        ["Cách viết", "Ưu tiên 3-5 câu: diễn biến chính, số liệu chỉ số/thanh khoản, cổ phiếu/nhóm ngành, bối cảnh."],
    ]
    for row in guide_rows:
        guide.append(row)
    guide.column_dimensions["A"].width = 24
    guide.column_dimensions["B"].width = 110
    guide["A1"].font = Font(bold=True)

    stats = wb.create_sheet("Pilot_Stats")
    stats.append(["Metric", "Value"])
    stats.append(["Total rows", len(rows)])
    for difficulty in ["Easy", "Medium", "Hard", "Extreme"]:
        stats.append([f"{difficulty} rows", sum(1 for row in rows if row["difficulty"] == difficulty)])
    stats.append(["Avg source words", round(sum(row["source_word_count"] for row in rows) / len(rows), 2) if rows else 0])
    stats.append(["Avg number count", round(sum(row["number_count"] for row in rows) / len(rows), 2) if rows else 0])
    for cell in stats[1]:
        cell.fill = header_fill
        cell.font = header_font
    stats.column_dimensions["A"].width = 28
    stats.column_dimensions["B"].width = 18

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="data/processed/annotation_sheet.csv")
    parser.add_argument("--output", default="data/processed/pilot_annotation.xlsx")
    parser.add_argument("--stats-output", default="data/processed/pilot_annotation_stats.json")
    parser.add_argument("--target", type=int, default=100)
    parser.add_argument("--allow-broad-finance", action="store_true")
    args = parser.parse_args()

    rows = prepare_rows(
        read_csv(Path(args.input)),
        target=args.target,
        require_stock_keyword=not args.allow_broad_finance,
    )
    write_workbook(rows, Path(args.output))

    stats = {
        "output": args.output,
        "rows": len(rows),
        "by_difficulty": {difficulty: sum(1 for row in rows if row["difficulty"] == difficulty) for difficulty in ["Easy", "Medium", "Hard", "Extreme"]},
        "avg_source_words": round(sum(row["source_word_count"] for row in rows) / len(rows), 2) if rows else 0,
        "avg_number_count": round(sum(row["number_count"] for row in rows) / len(rows), 2) if rows else 0,
    }
    Path(args.stats_output).write_text(json.dumps(stats, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(stats, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
