from __future__ import annotations

import argparse
import json
import math
import re
from pathlib import Path

from openpyxl import load_workbook

from visum.numbers import extract_number_mentions, extract_numbers
from visum.text import join_nonempty, sentence_split, word_count


KEYWORD_RE = re.compile(
    r"(VN-?Index|HNX-?Index|UPCoM|VN30|HoSE|HNX|chung khoan|chứng khoán|"
    r"co phieu|cổ phiếu|thanh khoan|thanh khoản|khoi ngoai|khối ngoại|"
    r"mua rong|mua ròng|ban rong|bán ròng|tu doanh|tự doanh|"
    r"doanh thu|loi nhuan|lợi nhuận|lai|lãi|lo|lỗ|co tuc|cổ tức|"
    r"tang|tăng|giam|giảm|diem|điểm|%)",
    flags=re.IGNORECASE,
)


def split_cell_numbers(value: object) -> list[str]:
    text = "" if value is None else str(value)
    return [item.strip() for item in text.replace("\n", ";").split(";") if item.strip()]


def norm(value: str) -> str:
    return value.lower().replace(" ", "")


def number_satisfies(required_norm: str, summary_norm: str) -> bool:
    if required_norm == summary_norm:
        return True
    if summary_norm.startswith(required_norm):
        suffix = summary_norm[len(required_norm) :]
        return bool(suffix) and not suffix[0].isdigit()
    return False


def number_mentions_cover(required: str, mentions: list[str]) -> bool:
    mention_norms = [norm(item) for item in mentions]
    required_parts = extract_numbers(required) or [required]
    required_norms = [norm(item) for item in required_parts]
    return all(
        any(number_satisfies(required_norm, mention_norm) for mention_norm in mention_norms)
        for required_norm in required_norms
    )


def covered_required_indices(required_numbers: list[str], text: str) -> set[int]:
    mentions = extract_numbers(text)
    return {
        index
        for index, required in enumerate(required_numbers)
        if number_mentions_cover(required, mentions)
    }


def strip_number_mentions(text: str) -> str:
    spans = [(item.start, item.end) for item in extract_number_mentions(text)]
    if not spans:
        return text

    output = text
    for start, end in sorted(spans, reverse=True):
        output = output[:start] + " " + output[end:]
    output = re.sub(r"\(\s*\)", " ", output)
    output = re.sub(r"\s+([,.;:!?])", r"\1", output)
    output = re.sub(r"\s+", " ", output).strip()
    return output


def sentence_score(sentence: str, index: int, required_numbers: list[str]) -> float:
    covered = covered_required_indices(required_numbers, sentence)
    number_bonus = min(len(extract_numbers(sentence)), 8) * 2.0
    keyword_bonus = 7.0 if KEYWORD_RE.search(sentence) else 0.0
    lead_bonus = max(0.0, 8.0 - index) if index < 8 else 0.0
    length_penalty = max(0, word_count(sentence) - 55) * 0.15
    return len(covered) * 60.0 + number_bonus + keyword_bonus + lead_bonus - length_penalty


def selected_sentence_text(selected: set[int], sentences: list[str]) -> str:
    return " ".join(sentences[index] for index in sorted(selected))


def add_sentence_if_fits(
    selected: set[int],
    sentences: list[str],
    candidate_index: int,
    max_words: int,
    tail: str = "",
) -> bool:
    trial = join_nonempty(
        [" ".join(sentences[index] for index in sorted(selected | {candidate_index})), tail]
    )
    if word_count(trial) <= max_words:
        selected.add(candidate_index)
        return True
    return False


def make_required_tail(summary: str, missing: list[str], max_words: int) -> tuple[str, list[str]]:
    if not missing:
        return "", []

    for end in range(len(missing), 0, -1):
        tail = "Số liệu cần giữ: " + "; ".join(missing[:end]) + "."
        trial = join_nonempty([summary, tail])
        if word_count(trial) <= max_words:
            return tail, missing[:end]
    return "", []


def draft_summary(
    text: str,
    required_numbers: list[str],
    min_words: int,
    max_words: int,
) -> tuple[str, dict]:
    sentences = sentence_split(text)
    sentences = [sentence for sentence in sentences if word_count(sentence) >= 5]
    if not sentences:
        return "", {"covered_required": 0, "required_total": len(required_numbers), "tail_added": 0}

    tail, tail_added = make_required_tail("", required_numbers, max_words)
    if len(tail_added) == len(required_numbers):
        scored_for_tail = [
            (sentence_score(sentence, index, required_numbers), index)
            for index, sentence in enumerate(sentences)
        ]
        selected_indices: set[int] = set()
        filler_parts: list[str] = []
        summary = tail

        if word_count(summary) < min_words:
            for _, index in sorted(scored_for_tail, key=lambda item: (-item[0], item[1])):
                if index in selected_indices:
                    continue
                candidate = strip_number_mentions(sentences[index])
                if word_count(candidate) < 5 or extract_numbers(candidate):
                    continue
                trial = join_nonempty([*filler_parts, candidate, tail])
                if word_count(trial) > max_words:
                    continue
                selected_indices.add(index)
                filler_parts.append(candidate)
                summary = trial
                if word_count(summary) >= min_words:
                    break

        covered_final = covered_required_indices(required_numbers, summary)
        return summary, {
            "covered_required": len(covered_final),
            "required_total": len(required_numbers),
            "tail_added": len(tail_added),
            "summary_words": word_count(summary),
        }

    selected: set[int] = set()

    if word_count(sentences[0]) <= max_words:
        selected.add(0)

    scored = [
        (sentence_score(sentence, index, required_numbers), index)
        for index, sentence in enumerate(sentences)
    ]

    while True:
        current = selected_sentence_text(selected, sentences)
        covered_now = covered_required_indices(required_numbers, current)
        best: tuple[float, int] | None = None

        for score, index in scored:
            if index in selected:
                continue
            new_cover = covered_required_indices(required_numbers, sentences[index]) - covered_now
            if not new_cover:
                continue
            trial_indices = sorted(selected | {index})
            trial = " ".join(sentences[item] for item in trial_indices)
            if word_count(trial) > max_words:
                continue
            gain_score = len(new_cover) * 1000.0 + score - index * 0.01
            if best is None or gain_score > best[0]:
                best = (gain_score, index)

        if best is None:
            break
        selected.add(best[1])

    summary = selected_sentence_text(selected, sentences)
    missing = [
        required
        for index, required in enumerate(required_numbers)
        if index not in covered_required_indices(required_numbers, summary)
    ]
    tail, tail_added = make_required_tail(summary, missing, max_words)
    summary = join_nonempty([summary, tail])

    if word_count(summary) < min_words:
        for _, index in sorted(scored, key=lambda item: (-item[0], item[1])):
            if index in selected:
                continue
            if add_sentence_if_fits(selected, sentences, index, max_words, tail=tail):
                summary = join_nonempty([selected_sentence_text(selected, sentences), tail])
                if word_count(summary) >= min_words:
                    break

    covered_final = covered_required_indices(required_numbers, summary)
    stats = {
        "covered_required": len(covered_final),
        "required_total": len(required_numbers),
        "tail_added": len(tail_added),
        "summary_words": word_count(summary),
    }
    return summary, stats


def build_drafts(input_path: Path, output_path: Path, stats_path: Path, overwrite: bool) -> dict:
    wb = load_workbook(input_path)
    ws = wb["Annotate"]
    headers = [cell.value for cell in ws[1]]
    col = {name: index + 1 for index, name in enumerate(headers)}

    total = 0
    drafted = 0
    skipped_filled = 0
    no_required = 0
    draft_word_counts: list[int] = []
    required_coverages: list[float] = []

    for row_index in range(2, ws.max_row + 1):
        total += 1
        text = str(ws.cell(row_index, col["text"]).value or "").strip()
        existing_summary = str(ws.cell(row_index, col["summary"]).value or "").strip()
        if existing_summary and not overwrite:
            skipped_filled += 1
            continue

        required = split_cell_numbers(ws.cell(row_index, col["required_numbers"]).value)
        if not required:
            no_required += 1
            continue

        source_words = word_count(text)
        min_words = max(
            int(ws.cell(row_index, col["min_summary_words"]).value or 0),
            math.ceil(source_words * 0.20),
        )
        max_words = min(
            int(ws.cell(row_index, col["max_summary_words"]).value or source_words),
            math.floor(source_words * 0.25),
        )
        summary, item_stats = draft_summary(text, required, min_words, max_words)
        if not summary:
            continue

        ws.cell(row_index, col["summary"]).value = summary
        ws.cell(row_index, col["validation_mode"]).value = "required_numbers"
        ws.cell(row_index, col["annotator_note"]).value = (
            "AUTO_DRAFT: review thu cong truoc khi dua vao dataset chinh thuc."
        )
        ws.cell(row_index, col["status"]).value = "Revise"

        drafted += 1
        draft_word_counts.append(item_stats["summary_words"])
        if item_stats["required_total"]:
            required_coverages.append(item_stats["covered_required"] / item_stats["required_total"])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    stats = {
        "input": str(input_path),
        "output": str(output_path),
        "total_rows": total,
        "drafted_rows": drafted,
        "skipped_filled_rows": skipped_filled,
        "rows_without_required_numbers": no_required,
        "avg_draft_words": round(sum(draft_word_counts) / len(draft_word_counts), 2) if draft_word_counts else 0,
        "avg_required_coverage_before_validation": round(sum(required_coverages) / len(required_coverages), 4)
        if required_coverages
        else 0,
    }
    stats_path.parent.mkdir(parents=True, exist_ok=True)
    stats_path.write_text(json.dumps(stats, ensure_ascii=False, indent=2), encoding="utf-8")
    return stats


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="data/processed/pilot_annotation_stock_batch2.xlsx")
    parser.add_argument("--output", default="data/processed/pilot_annotation_stock_batch2_draft.xlsx")
    parser.add_argument("--stats-output", default="data/processed/pilot_stock_batch2_draft_stats.json")
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args()

    stats = build_drafts(Path(args.input), Path(args.output), Path(args.stats_output), args.overwrite)
    print(json.dumps(stats, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
