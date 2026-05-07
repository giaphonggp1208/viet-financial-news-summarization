from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


def read_csv(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        return list(csv.DictReader(file))


def urls_from_csv(path: Path) -> set[str]:
    if not path.exists():
        return set()
    return {row.get("url", "") for row in read_csv(path) if row.get("url")}


def urls_from_excel(path: Path, sheet_name: str = "Annotate") -> set[str]:
    if not path.exists():
        return set()
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet_name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        url_index = headers.index("url")
    except ValueError:
        return set()
    urls = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        url = row[url_index]
        if url:
            urls.add(str(url))
    return urls


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--exclude-csv", action="append", default=[])
    parser.add_argument("--exclude-excel", action="append", default=[])
    args = parser.parse_args()

    excluded: set[str] = set()
    for path in args.exclude_csv:
        excluded.update(urls_from_csv(Path(path)))
    for path in args.exclude_excel:
        excluded.update(urls_from_excel(Path(path)))

    rows = [row for row in read_csv(Path(args.input)) if row.get("url", "") not in excluded]
    input_rows = read_csv(Path(args.input))
    fieldnames = list(input_rows[0].keys()) if input_rows else []

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"Input rows: {len(input_rows)}")
    print(f"Excluded URLs: {len(excluded)}")
    print(f"Saved rows: {len(rows)} to {out_path}")


if __name__ == "__main__":
    main()

