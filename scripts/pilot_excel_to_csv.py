from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="data/processed/pilot_annotation_validated.xlsx")
    parser.add_argument("--output", default="data/processed/annotation_from_pilot.csv")
    parser.add_argument("--status", choices=["Valid", "Revise", "Reject", "Not Started", "any"], default="Valid")
    args = parser.parse_args()

    wb = load_workbook(args.input, read_only=True)
    ws = wb["Annotate"]
    headers = [cell.value for cell in ws[1]]
    col = {name: index for index, name in enumerate(headers)}

    out_fields = ["id", "url", "source", "title", "published_at", "text", "summary", "notes"]
    rows: list[dict] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        status = values[col["status"]] or ""
        summary = values[col["summary"]] or ""
        if args.status != "any" and status != args.status:
            continue
        if not str(summary).strip():
            continue
        rows.append(
            {
                "id": values[col["id"]] or "",
                "url": values[col["url"]] or "",
                "source": values[col["source"]] or "",
                "title": values[col["title"]] or "",
                "published_at": values[col["published_at"]] or "",
                "text": values[col["text"]] or "",
                "summary": summary,
                "notes": f"exported_from={Path(args.input).name}; status={status}",
            }
        )

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=out_fields)
        writer.writeheader()
        writer.writerows(rows)

    print(f"Exported {len(rows)} rows to {out_path}")


if __name__ == "__main__":
    main()

