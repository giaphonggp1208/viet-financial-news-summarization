from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook

from visum.numbers import extract_numbers, extra_numbers, missing_numbers
from visum.text import compression_ratio, word_count


def split_numbers(cell_value: object) -> list[str]:
    text = "" if cell_value is None else str(cell_value)
    return [item.strip() for item in text.replace("\n", ";").split(";") if item.strip()]


def norm(value: str) -> str:
    return value.lower().replace(" ", "")


def number_satisfies(required_norm: str, summary_norm: str) -> bool:
    if required_norm == summary_norm:
        return True
    if summary_norm.startswith(required_norm):
        suffix = summary_norm[len(required_norm) :]
        return bool(suffix) and not suffix[0].isdigit()
    return False


def missing_required(required: list[str], summary: str) -> list[str]:
    summary_numbers = [norm(item) for item in extract_numbers(summary)]
    missing = []
    for item in required:
        required_parts = extract_numbers(item) or [item]
        required_norms = [norm(part) for part in required_parts]
        if not all(
            any(number_satisfies(required_norm, summary_norm) for summary_norm in summary_numbers)
            for required_norm in required_norms
        ):
            missing.append(item)
    return missing


def validate_workbook(input_path: Path, output_path: Path, stats_path: Path) -> dict:
    wb = load_workbook(input_path)
    ws = wb["Annotate"]
    headers = [cell.value for cell in ws[1]]
    col = {name: index + 1 for index, name in enumerate(headers)}

    number_ok_values = []
    length_ok_values = []
    ratios = []
    total = 0
    started = 0
    valid = 0
    revise = 0
    reject = 0

    for row_index in range(2, ws.max_row + 1):
        total += 1
        text = ws.cell(row_index, col["text"]).value or ""
        summary = ws.cell(row_index, col["summary"]).value or ""
        mode = str(ws.cell(row_index, col["validation_mode"]).value or "all_numbers").strip()
        source_words = word_count(text)
        summary_words = word_count(summary)
        ratio = compression_ratio(text, summary)
        length_ok = 0.20 <= ratio <= 0.25

        if summary.strip():
            started += 1

        if mode.startswith("required_numbers"):
            required = split_numbers(ws.cell(row_index, col["required_numbers"]).value)
            missing = missing_required(required, summary)
            extra = extra_numbers(text, summary)
            number_ok = not missing and not extra
        else:
            missing = missing_numbers(text, summary)
            extra = extra_numbers(text, summary)
            number_ok = not missing and not extra

        if not summary.strip():
            status = "Not Started"
        elif length_ok and number_ok:
            status = "Valid"
            valid += 1
        else:
            status = "Revise"
            revise += 1

        existing_status = ws.cell(row_index, col["status"]).value
        if existing_status == "Reject":
            status = "Reject"
            reject += 1

        ws.cell(row_index, col["summary_word_count"]).value = summary_words
        ws.cell(row_index, col["length_ratio"]).value = round(ratio, 4)
        ws.cell(row_index, col["length_compliance"]).value = "Pass" if length_ok else "Fail"
        ws.cell(row_index, col["missing_numbers"]).value = "; ".join(missing)
        ws.cell(row_index, col["hallucinated_numbers"]).value = "; ".join(extra)
        ws.cell(row_index, col["number_accuracy"]).value = 1.0 if number_ok else 0.0
        ws.cell(row_index, col["status"]).value = status

        if summary.strip() and status != "Reject":
            number_ok_values.append(1.0 if number_ok else 0.0)
            length_ok_values.append(1.0 if length_ok else 0.0)
            ratios.append(ratio)

    if "Validation_Summary" in wb.sheetnames:
        del wb["Validation_Summary"]
    summary_ws = wb.create_sheet("Validation_Summary")
    stats = {
        "total_rows": total,
        "started_rows": started,
        "valid_rows": valid,
        "revise_rows": revise,
        "reject_rows": reject,
        "quality_started_non_reject": {
            "number_accuracy": sum(number_ok_values) / len(number_ok_values) if number_ok_values else 0.0,
            "length_compliance": sum(length_ok_values) / len(length_ok_values) if length_ok_values else 0.0,
            "avg_compression_ratio": sum(ratios) / len(ratios) if ratios else 0.0,
            "n_samples": len(number_ok_values),
        },
    }
    for key, value in stats.items():
        summary_ws.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, dict) else value])
    summary_ws.column_dimensions["A"].width = 32
    summary_ws.column_dimensions["B"].width = 60

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    stats_path.write_text(json.dumps(stats, ensure_ascii=False, indent=2), encoding="utf-8")
    return stats


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="data/processed/pilot_annotation.xlsx")
    parser.add_argument("--output", default="data/processed/pilot_annotation_validated.xlsx")
    parser.add_argument("--stats-output", default="data/processed/pilot_validation_stats.json")
    args = parser.parse_args()

    stats = validate_workbook(Path(args.input), Path(args.output), Path(args.stats_output))
    print(json.dumps(stats, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
